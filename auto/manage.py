# coding=utf-8
from __future__ import division
from bs4 import BeautifulSoup
import xlsxwriter
from jtl import create_parser
from pandas import Series, DataFrame, ExcelWriter
from openpyxl import load_workbook
import pandas as pd
import re
import sys
import getopt
import os
import errno
import subprocess
import time
from string import Template
import platform
from fabric.api import *
from fabric.tasks import Task
import socket
import threading
import multiprocessing

_this_dir = os.path.dirname(os.path.abspath(__file__))


JMETER_Home = "D:\\apache-jmeter-3.3\\bin\\jmeter.bat"

env.hosts = ['10.3.6.15']
env.user = 'root'  # 多台主机用户名密码相同可以只写一次
env.password = '123.Ctrip'
env.warn_only = True

reload(sys)
sys.setdefaultencoding('utf8')


class BaseFunc(object):
    @staticmethod
    def execcmd(command):
        print "command " + command
        output = subprocess.Popen(
            command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True,
            universal_newlines=True)
        stderrinfo, stdoutinfo = output.communicate()
        # print "stderrinfo " + stderrinfo
        # print "stdoutinfo " + stdoutinfo
        # print "returncode={0}".format(output.returncode)
        return output.returncode

    def server_start(self, cmd):
        returnCode = self.execcmd(command=cmd)
        if returnCode == 1:
            print u"8787 Address already in use"

# 监控
class Monitor(object):
    def __init__(self, remote_ip, end_time, port=4, ):
        self.remote_ip = remote_ip
        self.port = port
        self.end_time = end_time

    # remote_ip = ['127.0.0.1', '10.33.20.20']
    # port = 4444

    def client(self, each_ip, port, end_time):
        print 'startMonitor'
        filename = each_ip + '.csv'
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        except socket.error, msg:
            print 'Failed to create socket. Error code: ' + str(msg[0]) + ' , Error message : ' + msg[1]
            sys.exit()

        try:
            s.connect((each_ip, port))
        except socket.gaierror:
            # could not resolve
            print 'Hostname could not be resolved. Exiting'
            sys.exit()
        try:
            message = 'test\ninterval:1\nmetrics:cpu:	memory:	\n'
            s.sendall(message)
        except socket.error:
            print 'Send failed'
            sys.exit()
        while time.time() <= end_time:
            with open(filename, 'a') as f:
                reply = s.recv(1024)
                reply = reply.replace('\t', ',')
                f.write(reply)
                f.flush()
                print reply
        end_mes = 'exit\n'
        s.sendall(end_mes)

    def start(self):
        threads = []
        for each_ip in self.remote_ip:
            threads.append(threading.Thread(target=self.client, args=(each_ip, self.port)))
        for t in threads:
            t.setDaemon(True)
            t.start()
        print "all over"


# 发布监控
class Fab(Task):
    name = 'deploy'

    def run(self, *args, **kwargs):
        self.task_upload()
        self.task_exc()

    # 打包
    @runs_once  # 该装饰器表示只执行一次，没有的话默认每台主机都执行一次
    def task_tar(self):  # 该场景本地文件打包本身就只需要执行一次
        with lcd('/home/python/test'):
            local('tar zcvf test.tar.gz test.py')

    # 上传
    def task_upload(self):
        run('mkdir -p /root/Desktop/ff')
        put('ServerAgent-2.2.3.zip', '/root/Desktop/ff/ServerAgent-2.2.3.zip')

    # 验证md5
    def task_md5(self):
        # 计算本地的md5
        local_md5 = local('md5sum /home/python/test/test.tar.gz', capture=True).split('  ')[0]
        # 计算远程主机md5
        remote_md5 = run('md5sum /home/python/temp/test.tar.gz').split('  ')[0]
        print(local_md5)
        print(remote_md5)
        if remote_md5 == local_md5:
            print('上传成功')
        else:
            print('上传出错')

    # 解包并执行
    def task_exc(self):
        with cd('/root/Desktop/ff'):
            run('unzip ServerAgent-2.2.3.zip')
            run('rm -rf ServerAgent-2.2.3.zip')
            run('(nohup sh /root/Desktop/ff/ServerAgent-2.2.3/startAgent.sh &) && sleep 1')
            time.sleep(5)
            run('pkill -f CMDRunner.jar')
            run('rm -rf ServerAgent-2.2.3')


instance = Fab()


# execcmd('fab -f manage.py start')

# 执行jmeter
class runJmeter(object):
    path = 'PerformanceTest'
    @staticmethod
    def getDateTime():
        '''
        获取当前日期时间，格式'08085159'
        '''
        return time.strftime(r'%m%d-%H%M%S', time.localtime(time.time()))

    @staticmethod
    def execjmxs(JmxFileName, Num_Threads, duration=600):
        print u'开始执行jmeter'
        runJmeter.mkdir()
        with open(JmxFileName, "r") as f:
            tmpstr = Template(f.read()).safe_substitute(
                num_threads=Num_Threads,
                duration=duration
            )
        now = runJmeter.getDateTime()
        tmpjmxfile = os.path.join(_this_dir, path) + "/script/{0}U_{1}.jmx".format(Num_Threads, now)
        with open(tmpjmxfile, "w+") as file:
            file.writelines(tmpstr)
        csvfilename = os.path.join(_this_dir, path) + "/result/{0}.csv".format(now)
        htmlreportpath = os.path.join(_this_dir, path) + "/result/report_{0}".format(now)
        if not os.path.exists(htmlreportpath):
            os.makedirs(htmlreportpath)
        if platform.system().lower() == 'windows':
            execjmxouthtml = "cmd.exe /c {JMETER_Home} -n -t {tmpjmxfile} -l {csvfilename} -e -o {htmlreportpath}" \
                .format(JMETER_Home=JMETER_Home, tmpjmxfile=tmpjmxfile, csvfilename=csvfilename,
                        htmlreportpath=htmlreportpath)
        else:
            execjmxouthtml = "{JMETER_Home} -n -t {tmpjmxfile} -l {csvfilename} -e -o {htmlreportpath}" \
                .format(JMETER_Home=JMETER_Home, tmpjmxfile=tmpjmxfile, csvfilename=csvfilename,
                        htmlreportpath=htmlreportpath)
        BaseFunc.execcmd(execjmxouthtml)

    @staticmethod
    def mkdir():
        currpath = os.path.dirname(os.path.realpath(__file__))
        folders = ['script', 'result']
        for each in folders:
            try:
                os.makedirs(os.path.join(currpath, path, each))
            except OSError as e:
                if e.errno == errno.EEXIST and os.path.isdir(path):
                    pass
                else:
                    raise


# 获取jmeter脚本信息
class GetInfo(object):
    def get_script(self):
        jmx_file = []
        script_jmx = ''
        _this_dir = os.path.dirname(os.path.abspath(__file__))
        l = os.listdir(_this_dir)
        for i in l:
            if os.path.splitext(i)[1] == '.jmx':
                jmx_file.append(i)
        print jmx_file
        if len(jmx_file) > 1:
            i = input(u"choose script(1,2,3):")
            script_jmx = jmx_file[i - 1]
        else:
            script_jmx = jmx_file[0]
        print script_jmx
        return script_jmx

    def get_request(self):
        JmxFileName = self.get_script()
        info = {}
        fo = open(JmxFileName, "r")
        script = fo.read()
        soup = BeautifulSoup(script, "xml")
        try:
            ip = soup.HTTPSamplerProxy.find("stringProp", attrs={"name": "HTTPSampler.domain"}).text
            if ip == '':
                ip = soup.ConfigTestElement.find("stringProp", attrs={"name": "HTTPSampler.domain"}).text
            if ip[0] == '$':
                name = ip[2:][:-1]
                try:
                    ip = soup.collectionProp.find("elementProp", attrs={"name": name}).find("stringProp", attrs={
                        "name": "Argument.value"}).text
                except:
                    ip = soup.hashTree.find("Arguments", attrs={"testname": "User Defined Variables"}). \
                        find("elementProp", attrs={"name": name}).find("stringProp",
                                                                       attrs={"name": "Argument.value"}).text
            request_text = soup.HTTPSamplerProxy.find("stringProp", attrs={"name": "Argument.value"}).text
        except AttributeError:
            print u'Error:找不到ip和请求报文'
            print u'不监控资源利用率'
        else:
            print u"接口ip:" + ip
            print u"报文:" + request_text
            info = {'ip': ip, "request": request_text, "script": JmxFileName}
            return info


# 结果文件
class Parse(object):
    def __init__(self):
        pass

    def csv_parse(self):
        obj = pd.read_csv("10-overall-summary.csv")
        obj['start_time'] = obj['timeStamp'] - obj['elapsed']
        temp = obj.groupby(obj["label"])
        test_time = (temp['timeStamp'].max() - temp['start_time'].min()) / 1000
        throughout = temp['timeStamp'].count() / test_time
        error = temp['success'].count() - temp['success'].sum()
        result = temp['elapsed'].agg(['min', 'max', 'mean', 'count'])
        result = DataFrame(result, columns=['min', 'max', 'mean', 'count', 'throughout', 'error'])
        result['throughout'] = throughout
        result['error'] = error
        Parse.w2e(result, 'result.xlsx', 'temp')

    def jtl_parse(self):
        response_time = []
        time_stamp = []
        start_times = []
        error_count = []
        threads = []
        lables = []
        parser = create_parser('10-overall-summary.csv')
        for sample in parser.itersamples():
            response_time.append(Parse.t2s(str(sample.elapsed_time)))
            time_stamp.append(sample.timestamp)
            lables.append(sample.label)
            threads.append(sample.group_threads)
            start_times.append(sample.timestamp - sample.elapsed_time)
            error_count.append((1, 0)[sample.success == True])
        data = {"response_time": response_time, "time_stamp": time_stamp, "lables": lables, "start_time": start_times,
                "error": error_count}
        obj = pd.DataFrame(data)
        temp = obj.groupby("lables")
        error = temp['error'].sum()
        test_time = temp['time_stamp'].max() - temp['start_time'].min()
        throughout = temp['time_stamp'].count() / test_time.astype('timedelta64[ms]') * 1000
        result = temp['response_time'].agg(['min', 'max', 'mean', 'count'])
        result = DataFrame(result, columns=['min', 'max', 'mean', 'count', 'throughout', 'error'])
        result['throughout'] = throughout
        result['error'] = error
        print result

    @staticmethod
    def w2e(df, excelFile, sheetName):
        xlsxwriter.Workbook(excelFile)
        # f.to_csv("test.csv", index="false") # 写入csv
        # writer = pd.ExcelWriter("test.xlsx", engine='xlsxwriter') # 写入xlsx,但是会覆盖
        writer = pd.ExcelWriter(excelFile, engine='openpyxl')  # 这个引擎不会覆盖
        book = load_workbook(excelFile)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, sheet_name=sheetName)
        writer.save()

    @staticmethod
    def parse_time(dt):
        # 转换成时间数组
        try:
            dt.split('.')[1]
            timeArray = time.strptime(dt, "%Y-%m-%d %H:%M:%S.%f")
        except:
            timeArray = time.strptime(dt, "%Y-%m-%d %H:%M:%S")
        # 转换成时间戳
        timestamp = time.mktime(timeArray)
        return timestamp

    @staticmethod
    def t2s(t):
        x = re.split(':|\.', str(t))
        if len(x) == 4:
            time_stamp = (int(x[0]) * 60 * 60 + int(x[1]) * 60 + int(x[2]) + float('0.' + x[3])) * 1000
        else:
            time_stamp = (int(x[0]) * 60 * 60 + int(x[1]) * 60 + int(x[2])) * 1000
        return time_stamp


# 命令使用解释
def usage():
    print '###################################'
    print '-r 不需要加参数,执行性能测试'
    print '-h 帮助'
    print '-m 监控服务器资源(linux)，如果服务器在参数文件中，则需要在文件中配置，否则自动获取'
    print '-t 执行时长'
    print '-u 执行用户数,多用户:5,10,15'
    print '-s 启动本地服务'
    print '-d '
    print '-f 输出目录,默认是当前目录'
    print '-a 分析'
    print '###################################'


def do():
    opts, args = getopt.getopt(sys.argv[1:], "ahrmt:u:s:d:f:", ["version", "file"])
    output_dir = ""
    duration = ""
    user = []
    sever_o = False
    exec_test = False
    monitor_if = False
    analyse = False
    for op, value in opts:
        if op == "-r":
            exec_test = True
        elif op == "-t":
            duration = value
        elif op == "-u":
            for each in value.split(","):
                user.append(each)
            user.sort()
        elif op == "-h":
            usage()
        elif op == "-f":
            output_dir = value
        elif op == "-m":
            monitor_if = True
        elif op == "-s":
            print "server start port 8787"
            sever_o = True
        elif op == '-a':
            analyse = True
        elif op == "--version":
            print "version 1.0"
            sys.exit()

    print u'执行时长:' + duration
    print u'执行用户' + ''.join(user)

    if exec_test:
        info = GetInfo().get_request()
        start_time = time.time()
        end_time = start_time + int(duration)
        jmx_script = info["script"]
        for thread in user:
            p = multiprocessing.Pool(5)
            # 执行测试
            p.apply_async(runJmeter.execjmxs(jmx_script, int(thread), int(duration)))
            if sever_o:
                p.apply_async(BaseFunc().server_start("python -m SimpleHTTPServer 8787"))
            if monitor_if:
                # 部署性能监控服务
                p.apply_async(BaseFunc.execcmd, args=('fab -f manage.py deploy',))
                # 监控
                p.apply_async(Monitor(remote_ip=info['ip'], port=4444, end_time=end_time).start, )
            p.close()
            p.join()
            print str(thread) + u"测试结束"
    if analyse:
        Parse().csv_parse()


if __name__ == '__main__':
    do()


